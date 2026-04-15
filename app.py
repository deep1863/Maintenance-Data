from flask import Flask, render_template, request, jsonify, send_file
import sqlite3, os, json, io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), 'instance', 'machine_data.db')

def get_desktop_path():
    home = os.path.expanduser("~")
    if os.name == 'nt':
        return os.path.join(os.environ.get('USERPROFILE', home), 'Desktop')
    return os.path.join(home, 'Desktop')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS machines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        department TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS maintenance_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        machine_id INTEGER NOT NULL,
        sno INTEGER,
        entry_date TEXT,
        breakdown_details TEXT,
        action_taken TEXT,
        spares_used TEXT,
        nature_of_work TEXT,
        nature_of_bd TEXT,
        total_down_time TEXT,
        bd_cleared TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now')),
        FOREIGN KEY(machine_id) REFERENCES machines(id)
    )''')
    # Seed some machines
    machines = [
        ('Hydraulic Press #1', 'Production'),
        ('CNC Machine #1', 'Machining'),
        ('Conveyor Belt A', 'Material Handling'),
        ('Mixer 30T', 'Processing'),
        ('Classifier Panel', 'Control'),
    ]
    for m in machines:
        c.execute('INSERT OR IGNORE INTO machines (name, department) VALUES (?, ?)', m)
    conn.commit()
    conn.close()

# ── Routes ──────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/machines', methods=['GET'])
def get_machines():
    conn = get_db()
    machines = conn.execute('SELECT * FROM machines ORDER BY name').fetchall()
    conn.close()
    return jsonify([dict(m) for m in machines])

@app.route('/api/machines', methods=['POST'])
def add_machine():
    data = request.json
    if not data.get('name'):
        return jsonify({'error': 'Machine name required'}), 400
    conn = get_db()
    try:
        conn.execute('INSERT INTO machines (name, department) VALUES (?, ?)',
                     (data['name'].strip(), data.get('department', '').strip()))
        conn.commit()
        machine = conn.execute('SELECT * FROM machines WHERE name=?', (data['name'].strip(),)).fetchone()
        return jsonify(dict(machine)), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'Machine already exists'}), 409
    finally:
        conn.close()

@app.route('/api/machines/<int:mid>', methods=['DELETE'])
def delete_machine(mid):
    conn = get_db()
    conn.execute('DELETE FROM maintenance_log WHERE machine_id=?', (mid,))
    conn.execute('DELETE FROM machines WHERE id=?', (mid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/entries', methods=['GET'])
def get_entries():
    machine_id = request.args.get('machine_id')
    search = request.args.get('search', '')
    conn = get_db()
    if machine_id:
        rows = conn.execute('''
            SELECT ml.*, m.name as machine_name FROM maintenance_log ml
            JOIN machines m ON ml.machine_id = m.id
            WHERE ml.machine_id = ? ORDER BY ml.entry_date DESC, ml.sno
        ''', (machine_id,)).fetchall()
    elif search:
        like = f'%{search}%'
        rows = conn.execute('''
            SELECT ml.*, m.name as machine_name FROM maintenance_log ml
            JOIN machines m ON ml.machine_id = m.id
            WHERE ml.breakdown_details LIKE ? OR ml.action_taken LIKE ?
               OR ml.spares_used LIKE ? OR m.name LIKE ?
            ORDER BY ml.entry_date DESC
        ''', (like, like, like, like)).fetchall()
    else:
        rows = conn.execute('''
            SELECT ml.*, m.name as machine_name FROM maintenance_log ml
            JOIN machines m ON ml.machine_id = m.id
            ORDER BY ml.entry_date DESC, ml.sno
        ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/entries', methods=['POST'])
def add_entry():
    data = request.json
    if not data.get('machine_id'):
        return jsonify({'error': 'Machine required'}), 400
    if not data.get('entry_date'):
        return jsonify({'error': 'Date required'}), 400
    if not data.get('breakdown_details'):
        return jsonify({'error': 'Breakdown details required'}), 400
    # Duplicate check
    conn = get_db()
    dup = conn.execute('''SELECT id FROM maintenance_log
        WHERE machine_id=? AND entry_date=? AND breakdown_details=?''',
        (data['machine_id'], data['entry_date'], data['breakdown_details'])).fetchone()
    if dup:
        conn.close()
        return jsonify({'error': 'Duplicate entry detected for same machine, date, and breakdown details'}), 409
    # Auto sno per machine
    last = conn.execute('SELECT MAX(sno) as m FROM maintenance_log WHERE machine_id=?',
                        (data['machine_id'],)).fetchone()
    sno = (last['m'] or 0) + 1
    conn.execute('''INSERT INTO maintenance_log
        (machine_id, sno, entry_date, breakdown_details, action_taken,
         spares_used, nature_of_work, nature_of_bd, total_down_time, bd_cleared)
        VALUES (?,?,?,?,?,?,?,?,?,?)''',
        (data['machine_id'], sno, data['entry_date'], data['breakdown_details'],
         data.get('action_taken',''), data.get('spares_used',''),
         data.get('nature_of_work','BD'), data.get('nature_of_bd','New'),
         data.get('total_down_time',''), data.get('bd_cleared','Int')))
    conn.commit()
    entry = conn.execute('SELECT ml.*, m.name as machine_name FROM maintenance_log ml JOIN machines m ON ml.machine_id=m.id WHERE ml.id=last_insert_rowid()').fetchone()
    conn.close()
    return jsonify(dict(entry)), 201

@app.route('/api/entries/<int:eid>', methods=['PUT'])
def update_entry(eid):
    data = request.json
    conn = get_db()
    conn.execute('''UPDATE maintenance_log SET
        entry_date=?, breakdown_details=?, action_taken=?,
        spares_used=?, nature_of_work=?, nature_of_bd=?,
        total_down_time=?, bd_cleared=?, updated_at=datetime('now')
        WHERE id=?''',
        (data['entry_date'], data['breakdown_details'], data.get('action_taken',''),
         data.get('spares_used',''), data.get('nature_of_work','BD'),
         data.get('nature_of_bd','New'), data.get('total_down_time',''),
         data.get('bd_cleared','Int'), eid))
    conn.commit()
    entry = conn.execute('SELECT ml.*, m.name as machine_name FROM maintenance_log ml JOIN machines m ON ml.machine_id=m.id WHERE ml.id=?', (eid,)).fetchone()
    conn.close()
    return jsonify(dict(entry))

@app.route('/api/entries/<int:eid>', methods=['DELETE'])
def delete_entry(eid):
    conn = get_db()
    conn.execute('DELETE FROM maintenance_log WHERE id=?', (eid,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/dashboard')
def dashboard():
    conn = get_db()
    total_entries = conn.execute('SELECT COUNT(*) as c FROM maintenance_log').fetchone()['c']
    total_machines = conn.execute('SELECT COUNT(*) as c FROM machines').fetchone()['c']
    bd_count = conn.execute("SELECT COUNT(*) as c FROM maintenance_log WHERE nature_of_work='BD'").fetchone()['c']
    pm_count = conn.execute("SELECT COUNT(*) as c FROM maintenance_log WHERE nature_of_work='PM'").fetchone()['c']
    machine_stats = conn.execute('''
        SELECT m.name, COUNT(ml.id) as entry_count,
               SUM(CASE WHEN ml.nature_of_work='BD' THEN 1 ELSE 0 END) as bd_count
        FROM machines m LEFT JOIN maintenance_log ml ON m.id=ml.machine_id
        GROUP BY m.id ORDER BY entry_count DESC
    ''').fetchall()
    recent = conn.execute('''
        SELECT ml.*, m.name as machine_name FROM maintenance_log ml
        JOIN machines m ON ml.machine_id=m.id
        ORDER BY ml.created_at DESC LIMIT 5
    ''').fetchall()
    conn.close()
    return jsonify({
        'total_entries': total_entries,
        'total_machines': total_machines,
        'bd_count': bd_count,
        'pm_count': pm_count,
        'machine_stats': [dict(r) for r in machine_stats],
        'recent': [dict(r) for r in recent]
    })

def build_excel(rows, machine_name="All Machines"):
    wb = Workbook()
    ws = wb.active
    ws.title = machine_name[:31]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    headers = ['S.No', 'Date', 'Machine', 'Details Of Breakdown Or Preventive Maintenance',
               'Details Of Action Taken', 'Spares Used', 'Nature Of Work',
               'Nature Of B/D (Repeat/New)', 'Total Down Time', 'B/D Cleared Int./Ext.']
    widths = [6, 12, 20, 40, 40, 20, 14, 20, 14, 18]
    ws.row_dimensions[1].height = 40
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    alt_fill = PatternFill("solid", fgColor="EBF1FA")
    for i, row in enumerate(rows, 2):
        ws.row_dimensions[i].height = 30
        fill = alt_fill if i % 2 == 0 else None
        values = [
            row.get('sno',''), row.get('entry_date',''), row.get('machine_name',''),
            row.get('breakdown_details',''), row.get('action_taken',''),
            row.get('spares_used',''), row.get('nature_of_work',''),
            row.get('nature_of_bd',''), row.get('total_down_time',''),
            row.get('bd_cleared','')
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin
            cell.alignment = center if col in [1,7,8,9,10] else left
            if fill:
                cell.fill = fill
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return wb

@app.route('/api/export/machine/<int:mid>')
def export_machine(mid):
    conn = get_db()
    machine = conn.execute('SELECT * FROM machines WHERE id=?', (mid,)).fetchone()
    rows = conn.execute('''SELECT ml.*, m.name as machine_name FROM maintenance_log ml
        JOIN machines m ON ml.machine_id=m.id WHERE ml.machine_id=?
        ORDER BY ml.entry_date, ml.sno''', (mid,)).fetchall()
    conn.close()
    if not machine:
        return jsonify({'error': 'Machine not found'}), 404
    wb = build_excel([dict(r) for r in rows], machine['name'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{machine['name'].replace(' ','_')}_maintenance.xlsx"
    return send_file(buf, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True)

@app.route('/api/export/all')
def export_all():
    conn = get_db()
    rows = conn.execute('''SELECT ml.*, m.name as machine_name FROM maintenance_log ml
        JOIN machines m ON ml.machine_id=m.id ORDER BY m.name, ml.entry_date, ml.sno''').fetchall()
    machines = conn.execute('SELECT * FROM machines ORDER BY name').fetchall()
    conn.close()
    rows_list = [dict(r) for r in rows]
    wb = build_excel(rows_list, "All Machines")
    # Add per-machine sheets
    for m in machines:
        m_rows = [r for r in rows_list if r['machine_id'] == m['id']]
        if not m_rows:
            continue
        ws = wb.create_sheet(title=m['name'][:31])
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ['S.No', 'Date', 'Details Of Breakdown Or PM',
                   'Details Of Action Taken', 'Spares Used', 'Nature Of Work',
                   'Nature Of B/D', 'Total Down Time', 'B/D Cleared']
        widths = [6, 12, 42, 42, 22, 14, 18, 14, 16]
        ws.row_dimensions[1].height = 40
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = thin
            ws.column_dimensions[get_column_letter(col)].width = w
        alt_fill = PatternFill("solid", fgColor="EBF1FA")
        for i, row in enumerate(m_rows, 2):
            ws.row_dimensions[i].height = 30
            fill = alt_fill if i % 2 == 0 else None
            for col, val in enumerate([
                row.get('sno',''), row.get('entry_date',''),
                row.get('breakdown_details',''), row.get('action_taken',''),
                row.get('spares_used',''), row.get('nature_of_work',''),
                row.get('nature_of_bd',''), row.get('total_down_time',''),
                row.get('bd_cleared','')
            ], 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin
                cell.alignment = center if col in [1,6,7,8,9] else left
                if fill: cell.fill = fill
        ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='master_maintenance_log.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True)

@app.route('/api/desktop-path')
def desktop_path():
    return jsonify({'path': get_desktop_path()})

if __name__ == '__main__':
    init_db()
    print("\n" + "="*55)
    print("  Machine Data Management System")
    print("  Running at: http://0.0.0.0:5000")
    print("  LAN Access: http://<YOUR-IP>:5000")
    print("="*55 + "\n")
    app.run(host='0.0.0.0', port=5000, debug=False)
